import os
import re
import sys
from pathlib import Path
from typing import List, Tuple

from dotenv import load_dotenv
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook

FOLDER_MIME = "application/vnd.google-apps.folder"
PDF_MIME = "application/pdf"
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


def extract_drive_id(url_or_id: str) -> str:
    s = (url_or_id or "").strip()
    if not s:
        raise ValueError("DRIVE_URL ว่างหรือไม่ถูกตั้งค่า")

    if re.fullmatch(r"[A-Za-z0-9_-]{10,}", s):
        return s

    patterns = [
        r"drive\.google\.com/drive/folders/([A-Za-z0-9_-]+)",
        r"drive\.google\.com/file/d/([A-Za-z0-9_-]+)",
        r"[?&]id=([A-Za-z0-9_-]+)",
        r"folders/([A-Za-z0-9_-]+)",
        r"/d/([A-Za-z0-9_-]+)",
    ]
    for pat in patterns:
        m = re.search(pat, s)
        if m:
            return m.group(1)

    raise ValueError("แยก ID จากลิงก์ไม่ได้: ใช้ลิงก์ drive/folders/<id> หรือส่งเป็น id ตรง ๆ")


def list_children(drive, folder_id: str) -> List[dict]:
    children: List[dict] = []
    page_token = None
    while True:
        resp = drive.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="nextPageToken, files(id,name,mimeType)",
            pageToken=page_token,
            pageSize=1000,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            corpora="allDrives",
        ).execute()

        children.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break

    # โฟลเดอร์ขึ้นก่อน แล้วเรียงตามชื่อ
    children.sort(key=lambda x: (x.get("mimeType") != FOLDER_MIME, (x.get("name") or "").lower()))
    return children


def collect_rows_folder1_folder3_pdfid(drive, root_id: str) -> List[Tuple[str, str, str]]:
    """
    เดินโครงสร้าง: root -> folder1 -> folder2 -> folder3 -> (pdf 1 ไฟล์)
    แล้วคืนค่าเป็น list ของ (folder1_name, folder3_name, pdf_id)
    """
    rows: List[Tuple[str, str, str]] = []

    level1 = [x for x in list_children(drive, root_id) if x.get("mimeType") == FOLDER_MIME]
    for f1 in level1:
        level2 = [x for x in list_children(drive, f1["id"]) if x.get("mimeType") == FOLDER_MIME]
        for f2 in level2:
            level3 = [x for x in list_children(drive, f2["id"]) if x.get("mimeType") == FOLDER_MIME]
            for f3 in level3:
                items = list_children(drive, f3["id"])

                # เอาเฉพาะ PDF และตัดสิ่งที่ไม่ต้องการ
                pdfs = [
                    it for it in items
                    if it.get("mimeType") == PDF_MIME
                    and (it.get("name") or "").lower().endswith(".pdf")
                    and (it.get("name") or "") != ".DS_Store"
                ]

                if not pdfs:
                    continue

                pdf = pdfs[0]  # โฟลเดอร์มีแค่ 1 ไฟล์ตามที่คุณบอก
                rows.append((f1.get("name", ""), f3.get("name", ""), pdf["id"]))

    return rows


def write_rows_to_excel(project_root: Path, rows: List[Tuple[str, str, str]]) -> None:
    xlsx_path = project_root / "curriculum.xlsx"
    if not xlsx_path.exists():
        print(f"ไม่เจอไฟล์ Excel: {xlsx_path}")
        sys.exit(1)

    wb = load_workbook(xlsx_path)
    sheet_name = "file_id"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    # ล้างข้อมูลเก่าในช่วง A2:C...
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).value = None
        ws.cell(r, 2).value = None
        ws.cell(r, 3).value = None

    # เขียนข้อมูล เริ่ม A2
    start_row = 2
    for i, (folder1, folder3, pdf_id) in enumerate(rows):
        r = start_row + i
        ws.cell(r, 1).value = folder1
        ws.cell(r, 2).value = folder3
        ws.cell(r, 3).value = pdf_id

    wb.save(xlsx_path)
    print(f"เขียนข้อมูลแล้ว: {xlsx_path} (ชีต '{sheet_name}' เริ่มที่ A2) จำนวน {len(rows)} แถว")


def main():
    project_root = Path(__file__).resolve().parents[1]

    # โหลด .env ที่ root
    load_dotenv()

    drive_url = os.getenv("DRIVE_URL")
    root_id = extract_drive_id(drive_url)

    sa_path = "service_account.json"

    creds = service_account.Credentials.from_service_account_file(str(sa_path), scopes=SCOPES)
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)

    try:
        rows = collect_rows_folder1_folder3_pdfid(drive, root_id)
        write_rows_to_excel(project_root, rows)
    except HttpError as e:
        print("เรียก Drive API ไม่สำเร็จ:", e)
        print("เช็กว่าแชร์โฟลเดอร์ให้ client_email ของ service account แล้ว และเปิด Drive API แล้ว")
        sys.exit(2)


if __name__ == "__main__":
    main()
